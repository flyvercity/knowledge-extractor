from pathlib import Path
from openpyxl import load_workbook


def extract_xlsx(file_path: Path, temp_dir: Path) -> str:
    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    lines = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n## {sheet_name}\n")
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(cells)
        if not rows:
            continue
        # Normalize column count
        max_cols = max(len(r) for r in rows)
        rows = [r + [""] * (max_cols - len(r)) for r in rows]
        header = "| " + " | ".join(rows[0]) + " |"
        sep = "| " + " | ".join("---" for _ in rows[0]) + " |"
        body = "\n".join("| " + " | ".join(r) + " |" for r in rows[1:])
        lines.append(f"{header}\n{sep}\n{body}\n")

    wb.close()
    return "\n".join(lines)
